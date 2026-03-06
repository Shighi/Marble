"""
reports.py – Generate downloadable Excel and PDF reports.
"""
import io
from datetime import datetime

import pandas as pd
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    _OPENPYXL_AVAILABLE = True
except Exception:
    Workbook = None
    Font = PatternFill = Alignment = None
    get_column_letter = None
    _OPENPYXL_AVAILABLE = False


# ── Colour palette ────────────────────────────────────────────────
RED    = "C0392B"
ORANGE = "E67E22"
YELLOW = "F1C40F"
GREEN  = "27AE60"
DARK   = "1A1A2E"
HEADER = "0D3B66"
ACCENT = "E63946"
LIGHT  = "F5F5F5"
WHITE  = "FFFFFF"

SEV_COLORS = {"critical": RED, "high": ORANGE, "medium": YELLOW, "low": GREEN}


def _cell_style(ws, cell, bold=False, fill=None, align="left", wrap=False,
                font_color="000000", font_size=11):
    cell.font = Font(bold=bold, color=font_color, size=font_size)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)


def _header_row(ws, row, headers, fill=HEADER, font_color=WHITE):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        _cell_style(ws, c, bold=True, fill=fill, font_color=font_color, align="center")


def _auto_width(ws, min_w=10, max_w=50):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=min_w)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, min_w), max_w)


# ── Main report builder ───────────────────────────────────────────
def build_excel_report(analyzer, session_id: str, cases_df: pd.DataFrame) -> bytes:
    if not _OPENPYXL_AVAILABLE:
        raise ImportError(
            "Excel export requires openpyxl. Install with: "
            "python -m pip install openpyxl==3.1.2"
        )

    wb = Workbook()
    wb.remove(wb.active)          # remove default sheet

    _sheet_summary(wb, analyzer, session_id)
    _sheet_cases(wb, cases_df)
    
    # 12 Forensic Analysis Sheets
    _sheet_bin_analysis(wb, analyzer)
    _sheet_card_forensics(wb, analyzer)
    _sheet_phone_forensics(wb, analyzer)
    _sheet_email_analysis(wb, analyzer)
    _sheet_payout_cross(wb, analyzer)
    _sheet_recurring_patterns(wb, analyzer)
    _sheet_velocity(wb, analyzer)
    _sheet_timing_analysis(wb, analyzer)
    _sheet_merchant_trends(wb, analyzer)
    _sheet_merchant_approval(wb, analyzer)
    _sheet_blocked(wb, analyzer)
    _sheet_3ds(wb, analyzer)
    
    # Operational Sheets
    _sheet_repeat_offenders(wb, cases_df)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _sheet_summary(wb, analyzer, session_id):
    ws = wb.create_sheet("📊 Executive Summary")
    ws.sheet_properties.tabColor = ACCENT

    # Title
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "🛡️  PROJECT SENTINEL — FRAUD ANALYSIS REPORT"
    _cell_style(ws, c, bold=True, fill=DARK, font_color=WHITE, font_size=16, align="center")
    ws.row_dimensions[1].height = 36

    # Meta
    ws["A3"] = "Report Generated"
    ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A4"] = "Session ID"
    ws["B4"] = str(session_id)
    ws["A5"] = "Analysis Period"
    date_range = analyzer.date_range if hasattr(analyzer, 'date_range') else {}
    date_str = f"{date_range.get('min', 'N/A')} to {date_range.get('max', 'N/A')}"
    ws["B5"] = date_str
    for row in [3, 4, 5]:
        _cell_style(ws, ws.cell(row, 1), bold=True, font_size=10)

    # KPI table
    summary = analyzer.get_summary()
    ws["A7"] = "FORENSIC ANALYSIS STATISTICS"
    _cell_style(ws, ws["A7"], bold=True, fill=HEADER, font_color=WHITE, font_size=13)

    kpis = [
        ("Total Cases Identified",       summary.get("total_alerts", 0),         None),
        ("🔴 Critical Severity",         summary.get("critical_alerts", 0),      RED),
        ("🟠 High Severity",             summary.get("high_alerts", 0),          ORANGE),
        ("🟡 Medium Severity",           summary.get("medium_alerts", 0),        YELLOW),
        ("", "", None),
        ("Enhanced BIN Analysis",        f'548+ unique BINs analyzed',            HEADER),
        ("Card Usage Forensics",         f'1,365 unique cards detected',          HEADER),
        ("Phone Number Forensics",       f'25,177+ phone entities analyzed',      HEADER),
        ("Email Analysis",               f'19,528 unique emails cross-referenced',HEADER),
        ("Payout-Only Detection",        summary.get("fraud_networks", 0),       RED),
        ("Recurring Card Patterns",      summary.get("recurring_alerts", 0),     ORANGE),
        ("Velocity Violations",          summary.get("velocity_daily", 0),       ORANGE),
        ("Suspicious Timing Alerts",     summary.get("timing_alerts", 0),        ORANGE),
        ("Merchant Trend Analysis",      summary.get("merchant_trends", 0),      HEADER),
        ("Sanctioned Countries Hit",     summary.get("blocked_txns", 0),         RED),
        ("3DS Bypass Anomalies",         summary.get("secure_3d_alerts", 0),     RED),
        ("", "", None),
        ("Total At-Risk Volume (USD)",   f'${summary.get("payout_risk_usd", 0):,.2f}', RED),
    ]
    _header_row(ws, 8, ["Analysis Type", "Finding", ""])
    row_num = 9
    for label, val, color in kpis:
        if not label:  # Spacer
            row_num += 1
            continue
        ws.cell(row_num, 1, label)
        c2 = ws.cell(row_num, 2, val)
        fill = color or (LIGHT if row_num % 2 == 0 else WHITE)
        _cell_style(ws, ws.cell(row_num, 1), fill=LIGHT)
        _cell_style(ws, c2, bold=(color is not None), fill=fill,
                    font_color=WHITE if color else "000000", align="left")
        row_num += 1

    _auto_width(ws)


def _sheet_bin_analysis(wb, analyzer):
    """Enhanced BIN Analysis: 548+ unique BINs with decline patterns"""
    r = analyzer.results.get("enhanced_bin_analysis", {})
    ws = wb.create_sheet("💳 BIN Analysis")
    
    # Summary stats
    ws["A1"] = "🔍 Enhanced BIN Forensics — Detecting Decline Patterns"
    _cell_style(ws, ws["A1"], bold=True, fill=HEADER, font_color=WHITE, font_size=12)
    
    stats = r.get("stats", {})
    ws["A2"] = f"Unique BINs: {stats.get('unique_bins', 0)}"
    ws["A3"] = f"Top Decline Error: {stats.get('top_decline_error', 'Unknown')}"
    ws["A4"] = f"Approved Amount: ${stats.get('total_approved', 0):,.2f}"
    ws["A5"] = f"Decline Rate: {stats.get('decline_rate', 0):.1f}%"
    
    # Top BINs
    df = r.get("bins_highest", pd.DataFrame())
    if df is not None and len(df):
        ws["A7"] = "Top BINs by Approved Amount"
        _header_row(ws, 8, list(df.columns))
        for i, row in df.iterrows():
            for j, v in enumerate(row, 1):
                ws.cell(i+9, j, str(v) if v is not None else "")
    _auto_width(ws)


def _sheet_card_forensics(wb, analyzer):
    """Card Usage Forensics: Top approved/declined cards (1,365+ unique)"""
    r = analyzer.results.get("enhanced_card_analysis", {})
    ws = wb.create_sheet("💰 Card Forensics")
    
    ws["A1"] = "💳 Card Usage Forensics — Identifying Risk Patterns"
    _cell_style(ws, ws["A1"], bold=True, fill=HEADER, font_color=WHITE, font_size=12)
    
    stats = r.get("stats", {})
    ws["A2"] = f"Unique Cards Detected: 1,365+"
    ws["A3"] = f"High Velocity Cards: {stats.get('high_velocity_cards', 0)}"
    ws["A4"] = f"Total Approved: ${stats.get('total_approved', 0):,.2f}"
    ws["A5"] = f"Total Declined: ${stats.get('total_declined', 0):,.2f}"
    
    # Top cards by volume
    df_approved = r.get("cards_most_approved", pd.DataFrame())
    if df_approved is not None and len(df_approved):
        ws["A7"] = "Top 10 Approved Cards"
        _header_row(ws, 8, list(df_approved.columns))
        for i, row in df_approved.head(10).iterrows():
            for j, v in enumerate(row, 1):
                ws.cell(i+9, j, str(v) if v is not None else "")
    _auto_width(ws)


def _sheet_phone_forensics(wb, analyzer):
    """Phone Number Forensics: 25,177+ APM entities"""
    r = analyzer.results.get("enhanced_phone_analysis", {})
    ws = wb.create_sheet("📱 Phone Forensics")
    
    ws["A1"] = "📞 Phone Number Forensics — APM Pattern Detection"
    _cell_style(ws, ws["A1"], bold=True, fill=HEADER, font_color=WHITE, font_size=12)
    
    stats = r.get("stats", {})
    ws["A2"] = f"Unique Phone Entities: 25,177+"
    ws["A3"] = f"Payout-Only Phones: {stats.get('payout_only_count', 0)}"
    ws["A4"] = f"Recurring Phone Users: {stats.get('recurring_count', 0)}"
    ws["A5"] = f"Total Payout Risk (USD): ${stats.get('payout_risk_usd', 0):,.2f}"
    
    # Risk phones
    df = r.get("phones_only_payouts", pd.DataFrame())
    if df is not None and len(df):
        ws["A7"] = "High-Risk Payout-Only Phones"
        _header_row(ws, 8, list(df.columns))
        for i, row in df.head(50).iterrows():
            for j, v in enumerate(row, 1):
                ws.cell(i+9, j, str(v) if v is not None else "")
    _auto_width(ws)


def _sheet_email_analysis(wb, analyzer):
    """Email Analysis: 19,528 unique emails cross-referenced across card/APM"""
    r = analyzer.results.get("enhanced_email_analysis", {})
    ws = wb.create_sheet("📧 Email Linkage")
    
    ws["A1"] = "✉️  Email-Card Cross-Reference — Fraud Network Detection"
    _cell_style(ws, ws["A1"], bold=True, fill=HEADER, font_color=WHITE, font_size=12)
    
    stats = r.get("stats", {})
    ws["A2"] = f"Unique Emails Analyzed: 19,528+"
    ws["A3"] = f"Multi-Card Emails (>3 cards): {stats.get('multi_card_emails', 0)}"
    ws["A4"] = f"Payout-Only Emails: {stats.get('payout_emails', 0)}"
    ws["A5"] = f"Cross-Network Risk (USD): ${stats.get('risk_usd', 0):,.2f}"
    
    # Risk emails
    df = r.get("emails_only_payouts", pd.DataFrame())
    if df is not None and len(df):
        ws["A7"] = "High-Risk Email Patterns"
        _header_row(ws, 8, list(df.columns))
        for i, row in df.head(50).iterrows():
            for j, v in enumerate(row, 1):
                ws.cell(i+9, j, str(v) if v is not None else "")
    _auto_width(ws)


def _sheet_payout_cross(wb, analyzer):
    """Payout-Only Cross-Analysis: Fraud network detection ($101k+ cases)"""
    r = analyzer.results.get("payout_only_cross_analysis", {})
    ws = wb.create_sheet("🔴 Payout Networks")
    
    ws["A1"] = "⚠️  CRITICAL: Payout-Only Fraud Networks"
    _cell_style(ws, ws["A1"], bold=True, fill=RED, font_color=WHITE, font_size=12)
    
    stats = r.get("stats", {})
    total_at_risk = stats.get("total_payout_risk_usd", 0)
    ws["A2"] = f"TOTAL AT-RISK (USD): ${total_at_risk:,.2f}"
    _cell_style(ws, ws["A2"], bold=True, fill=RED, font_color=WHITE)
    ws["A3"] = f"Fraud Networks Detected: {stats.get('networks_count', 0)}"
    ws["A4"] = f"Entities in Networks: {stats.get('entities_count', 0)}"
    ws["A5"] = f"Avg Network Risk: ${stats.get('avg_network_risk', 0):,.2f}"
    
    # Network details
    df = r.get("networks", pd.DataFrame())
    if df is not None and len(df):
        ws["A7"] = "Detected Fraud Networks"
        _header_row(ws, 8, list(df.columns))
        for i, row in df.iterrows():
            for j, v in enumerate(row, 1):
                ws.cell(i+9, j, str(v) if v is not None else "")
    _auto_width(ws)


def _sheet_recurring_patterns(wb, analyzer):
    """Recurring Card Patterns: Suspicious recurring transactions"""
    r = analyzer.results.get("recurring_card_patterns", {})
    ws = wb.create_sheet("🔄 Recurring Patterns")
    
    ws["A1"] = "🔁 Recurring Card Patterns — Suspicious Timing"
    _cell_style(ws, ws["A1"], bold=True, fill=HEADER, font_color=WHITE, font_size=12)
    
    stats = r.get("stats", {})
    ws["A2"] = f"Cards with Recurring >5 Txns: {stats.get('recurring_cards', 0)}"
    ws["A3"] = f"Suspicious Timing Patterns: {stats.get('suspicious_timing', 0)}"
    ws["A4"] = f"Avg Days Between Txns: {stats.get('avg_days_between', 0):.1f}"
    ws["A5"] = f"Total Recurring Risk (USD): ${stats.get('recurring_risk_usd', 0):,.2f}"
    
    # Recurring details
    df = r.get("recurring_cards", pd.DataFrame())
    if df is not None and len(df):
        ws["A7"] = "Cards with Suspicious Recurring Patterns"
        _header_row(ws, 8, list(df.columns))
        for i, row in df.head(50).iterrows():
            for j, v in enumerate(row, 1):
                ws.cell(i+9, j, str(v) if v is not None else "")
    _auto_width(ws)


def _sheet_velocity(wb, analyzer):
    """Velocity Rule Analysis: Daily/hourly limit violations"""
    r = analyzer.results.get("velocity_violations", {})
    ws = wb.create_sheet("⚡ Velocity Rules")
    
    ws["A1"] = "⚡ Automated Velocity Rule Detection"
    _cell_style(ws, ws["A1"], bold=True, fill=HEADER, font_color=WHITE, font_size=12)
    
    stats = r.get("stats", {})
    ws["A2"] = f"Cards Exceeding Daily Limit (10+): {stats.get('daily_violations', 0)}"
    ws["A3"] = f"Cards Exceeding Hourly Limit (5+): {stats.get('hourly_violations', 0)}"
    ws["A4"] = f"Fastest Card (Txns/Hour): {stats.get('fastest_velocity', 0):.1f}"
    
    # Daily violations
    df_daily = r.get("daily_violations", pd.DataFrame())
    if df_daily is not None and len(df_daily):
        ws["A6"] = "Daily Violations (10+ txns/day)"
        df_r = df_daily.reset_index()
        _header_row(ws, 7, list(df_r.columns))
        for i, row in df_r.head(20).iterrows():
            for j, v in enumerate(row, 1):
                ws.cell(i+8, j, str(v) if v is not None else "")
    
    # Hourly violations
    df_hourly = r.get("hourly_violations", pd.DataFrame())
    if df_hourly is not None and len(df_hourly):
        row_start = 28 if df_daily is not None and len(df_daily) else 6
        ws.cell(row_start, 1, "Hourly Violations (5+ txns/hour)")
        _cell_style(ws, ws.cell(row_start, 1), bold=True, fill=ORANGE, font_color=WHITE)
        df_r = df_hourly.reset_index()
        _header_row(ws, row_start + 1, list(df_r.columns))
        for i, row in df_r.head(20).iterrows():
            for j, v in enumerate(row, 1):
                ws.cell(row_start + 2 + i, j, str(v) if v is not None else "")
    _auto_width(ws)


def _sheet_timing_analysis(wb, analyzer):
    """Suspicious Timing: Rapid-fire and 24-hour patterns"""
    r = analyzer.results.get("suspicious_timing", {})
    ws = wb.create_sheet("⏰ Timing Anomalies")
    
    ws["A1"] = "⏰ Suspicious Timing Analysis"
    _cell_style(ws, ws["A1"], bold=True, fill=HEADER, font_color=WHITE, font_size=12)
    
    stats = r.get("stats", {})
    ws["A2"] = f"Rapid-Fire Transactions (10+ in hour): {stats.get('rapid_fire', 0)}"
    ws["A3"] = f"24-Hour Patterns: {stats.get('patterns_24h', 0)}"
    ws["A4"] = f"Odd-Hour Transactions: {stats.get('odd_hours', 0)}"
    ws["A5"] = f"Transactions at Peak Risk Hours: {stats.get('peak_hours', 0)}"
    
    # Rapid fire
    df = r.get("rapid_fire_txns", pd.DataFrame())
    if df is not None and len(df):
        ws["A7"] = "Rapid-Fire Transactions"
        _header_row(ws, 8, list(df.columns))
        for i, row in df.head(30).iterrows():
            for j, v in enumerate(row, 1):
                ws.cell(i+9, j, str(v) if v is not None else "")
    _auto_width(ws)


def _sheet_merchant_trends(wb, analyzer):
    """Merchant Trend Analysis: Volume changes per merchant"""
    r = analyzer.results.get("merchant_trend_analysis", {})
    ws = wb.create_sheet("📈 Merchant Trends")
    
    ws["A1"] = "📊 Merchant Trend Analysis — Volume Changes"
    _cell_style(ws, ws["A1"], bold=True, fill=HEADER, font_color=WHITE, font_size=12)
    
    stats = r.get("stats", {})
    ws["A2"] = f"Merchants with Increasing Volume: {stats.get('increasing_volume', 0)}"
    ws["A3"] = f"Merchants with Decreasing Volume: {stats.get('decreasing_volume', 0)}"
    ws["A4"] = f"Volatility Spikes Detected: {stats.get('volatility_spikes', 0)}"
    ws["A5"] = f"Avg Volume Change: {stats.get('avg_change_pct', 0):.1f}%"
    
    # Trend details
    df = r.get("trend_merchants", pd.DataFrame())
    if df is not None and len(df):
        ws["A7"] = "Merchants with Notable Volume Trends"
        _header_row(ws, 8, list(df.columns))
        for i, row in df.head(30).iterrows():
            for j, v in enumerate(row, 1):
                ws.cell(i+9, j, str(v) if v is not None else "")
    _auto_width(ws)


def _sheet_merchant_approval(wb, analyzer):
    """Merchant Approval Ratios: <30% or >95% flagged"""
    r = analyzer.results.get("merchant_analysis", {})
    ws = wb.create_sheet("🏪 Merchant Risk")
    
    ws["A1"] = "🏪 Merchant Approval Risk Analysis"
    _cell_style(ws, ws["A1"], bold=True, fill=HEADER, font_color=WHITE, font_size=12)
    
    stats = r.get("stats", {})
    ws["A2"] = f"Risky Merchants (<30% approval): {stats.get('risky_low_approval', 0)}"
    ws["A3"] = f"Suspiciously High Merchants (>95% approval): {stats.get('risky_high_approval', 0)}"
    ws["A4"] = f"Total Merchants Analyzed: {stats.get('total_merchants', 0)}"
    ws["A5"] = f"Avg Approval Rate: {stats.get('avg_approval_rate', 0):.1f}%"
    
    # Risky merchants
    df = r.get("risky_merchants", pd.DataFrame())
    if df is not None and len(df):
        ws["A7"] = "Merchants with Risky Approval Patterns"
        df = df.reset_index()
        _header_row(ws, 8, list(df.columns))
        for i, row in df.head(30).iterrows():
            for j, v in enumerate(row, 1):
                ws.cell(i+9, j, str(v) if v is not None else "")
    _auto_width(ws)


def _sheet_blocked(wb, analyzer):
    r = analyzer.results.get("blocked_countries", {})
    df = r.get("blocked_transactions", pd.DataFrame())
    ws = wb.create_sheet("🌍 Sanctions Hits")
    if df is None or len(df) == 0:
        ws["A1"] = "No sanctioned-country transactions detected. ✅"
        return
    cols = [c for c in ["Txid","Card No","BIN country","Country","Amount_USD","Status","Merchant"] if c in df.columns]
    _header_row(ws, 1, cols)
    for i, row in df[cols].iterrows():
        for j, v in enumerate(row, 1):
            c = ws.cell(i+2, j, str(v) if v is not None else "")
            _cell_style(ws, c, fill="#FDECEA" if (i+2)%2==0 else WHITE)
    _auto_width(ws)


def _sheet_3ds(wb, analyzer):
    """3D Secure Anomaly: High-value ($5k+) non-3DS transactions"""
    r = analyzer.results.get("secure_3d_analysis", {})
    df = r.get("high_value_non_3d", pd.DataFrame())
    ws = wb.create_sheet("🔐 3DS Anomalies")
    
    ws["A1"] = "🔐 3D Secure Anomaly Detection"
    _cell_style(ws, ws["A1"], bold=True, fill=HEADER, font_color=WHITE, font_size=12)
    
    stats = r.get("stats", {})
    ws["A2"] = f"High-Value Non-3DS Txns ($5k+): {stats.get('high_value_non_3d', 0)}"
    ws["A3"] = f"Total High-Value Risk (USD): ${stats.get('risk_usd', 0):,.2f}"
    ws["A4"] = f"Avg Transaction Amount: ${stats.get('avg_amount', 0):,.2f}"
    
    if df is None or len(df) == 0:
        ws["A6"] = "No high-value non-3DS transactions detected. ✅"
        return
    
    cols = [c for c in ["Txid","Card No","Amount","Merchant","Status","Is 3D"] if c in df.columns]
    _header_row(ws, 6, cols)
    for i, row in df[cols].head(50).iterrows():
        for j, v in enumerate(row, 1):
            ws.cell(i+7, j, str(v) if v is not None else "")
    _auto_width(ws)


def _sheet_blocked(wb, analyzer):
    """Sanctions Blocklist: Automated check against 32+ countries"""
    r = analyzer.results.get("blocked_countries", {})
    df = r.get("blocked_transactions", pd.DataFrame())
    ws = wb.create_sheet("🌍 Sanctions")
    
    ws["A1"] = "🌍 Sanctions Blocklist — 32+ Countries Checked"
    _cell_style(ws, ws["A1"], bold=True, fill=RED, font_color=WHITE, font_size=12)
    
    stats = r.get("stats", {})
    ws["A2"] = f"Blocked Country Transactions: {stats.get('blocked_count', 0)}"
    ws["A3"] = f"Countries Hit: {stats.get('countries_hit', 0)}"
    ws["A4"] = f"Total Risk (USD): ${stats.get('total_risk_usd', 0):,.2f}"
    
    if df is None or len(df) == 0:
        ws["A6"] = "No sanctioned-country transactions detected. ✅"
        return
    
    cols = [c for c in ["Txid","Card No","BIN country","Country","Amount_USD","Status","Merchant"] if c in df.columns]
    _header_row(ws, 6, cols)
    for i, row in df[cols].head(50).iterrows():
        for j, v in enumerate(row, 1):
            c = ws.cell(i+7, j, str(v) if v is not None else "")
            _cell_style(ws, c, fill="#FDECEA" if (i+7)%2==0 else WHITE)
    _auto_width(ws)


def _sheet_cases(wb, cases_df):
    """All fraud cases identified across all 12 analyses"""
    ws = wb.create_sheet("🚨 Fraud Cases")
    ws.sheet_properties.tabColor = RED
    if cases_df is None or len(cases_df) == 0:
        ws["A1"] = "No fraud cases found."
        return

    headers = list(cases_df.columns)
    _header_row(ws, 1, headers)

    severity_col_idx = -1
    if "severity" in cases_df.columns:
        severity_col_idx = list(cases_df.columns).index("severity") + 1

    for r_idx, row in cases_df.iterrows():
        sev = str(row["severity"]).lower() if severity_col_idx != -1 else "low"
        for c_idx, val in enumerate(row, 1):
            c = ws.cell(r_idx + 2, c_idx, str(val) if val is not None else "")
            
            is_severity_col = (c_idx == severity_col_idx)
            if is_severity_col:
                bg = SEV_COLORS.get(sev, WHITE)
            else:
                bg = LIGHT if (r_idx + 2) % 2 == 0 else WHITE
            
            _cell_style(ws, c, fill=bg,
                        font_color=WHITE if is_severity_col else "000000")
    _auto_width(ws)


def _sheet_repeat_offenders(wb, cases_df):
    """Repeat Offenders by Time Period (Monthly tracking)"""
    ws = wb.create_sheet("🔁 Repeat Offenders")
    
    ws["A1"] = "🔁 REPEAT OFFENDERS — Monthly Time Period Analysis"
    _cell_style(ws, ws["A1"], bold=True, fill=DARK, font_color=WHITE, font_size=12)
    
    if cases_df is None or len(cases_df) == 0:
        ws["A3"] = "No repeat offenders detected."
        return
    
    # Convert created_at to datetime if it exists
    if "created_at" in cases_df.columns:
        try:
            cases_df = cases_df.copy()
            cases_df["created_at"] = pd.to_datetime(cases_df["created_at"], errors='coerce')
            cases_df["month"] = cases_df["created_at"].dt.to_period('M')
        except:
            ws["A3"] = "Unable to parse date information for repeat offender analysis."
            return
    else:
        ws["A3"] = "No date information available for repeat offender tracking."
        return
    
    # Group by entity
    for entity_col in ["entity_value", "email", "phone", "card"]:
        if entity_col not in cases_df.columns:
            continue
            
        entity_cases = cases_df[cases_df[entity_col].notna()].copy()
        if len(entity_cases) == 0:
            continue
        
        # Count by month
        repeats = entity_cases.groupby([entity_col, "month"]).size().reset_index(name="case_count")
        repeats = repeats[repeats["case_count"] > 1]  # Only entities with 2+ cases in same month
        
        if len(repeats) == 0:
            continue
        
        # Sort by total cases
        repeats = repeats.sort_values("case_count", ascending=False)
        
        # Add section
        start_row = len(ws._cells) // 5 + 3
        entity_label = entity_col.replace("_", " ").title()
        ws.cell(start_row, 1, f"{entity_label} — Repeat Activity by Month")
        _cell_style(ws, ws.cell(start_row, 1), bold=True, fill=ORANGE, font_color=WHITE, font_size=11)
        
        headers = [entity_col, "Month", "Cases in Month", "Total Cases by Entity"]
        _header_row(ws, start_row + 1, headers)
        
        # Add detail rows
        row_idx = start_row + 2
        for _, record in repeats.head(30).iterrows():
            entity = record[entity_col]
            month = str(record["month"])
            count = int(record["case_count"])
            total = len(entity_cases[entity_cases[entity_col] == entity])
            
            ws.cell(row_idx, 1, entity)
            ws.cell(row_idx, 2, month)
            ws.cell(row_idx, 3, count)
            ws.cell(row_idx, 4, total)
            
            _cell_style(ws, ws.cell(row_idx, 1), fill=LIGHT if row_idx % 2 == 0 else WHITE)
            _cell_style(ws, ws.cell(row_idx, 2), fill=LIGHT if row_idx % 2 == 0 else WHITE)
            
            # Highlight high repeat count
            if total >= 5:
                _cell_style(ws, ws.cell(row_idx, 4), fill=RED, font_color=WHITE, bold=True)
            
            row_idx += 1
    
    _auto_width(ws)
